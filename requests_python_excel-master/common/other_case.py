# coding:utf-8
import requests
import json
from common import database
from openpyxl import load_workbook

# 取token并填入headers
url2 = 'http://student-manage-test.lxhelper.com/api/v1/user_info/login/'
body2 = {"username": "admin", "password": "fc76c4a86c56becc717a88f651264622"}
r2 = requests.post(url=url2, data=body2)
login_token = r2.headers['token']
header2 = {"Content-Type": "application/json", "token": login_token}


# 将headers填入excle
def excle_save():
    wb = load_workbook("E:\\git_zip\\interface-master\\requests_python_excel-master\\case\\demo_api.xlsx")
    wb_new = wb.active
    a = 1
    xl_sheet_names = wb.get_sheet_names()
    xl_sheet = wb.get_sheet_by_name(xl_sheet_names[0])
    row = xl_sheet.max_row
    while a < row:
        a += 1
        wb_new.cell(a, 5, str(header2))
        wb.save("E:\\git_zip\\interface-master\\requests_python_excel-master\\case\\demo_api.xlsx")


# 可以写其他需要参数传递的接口
def student_info():
    url = 'http://student-manage-test.lxhelper.com/api/v1/student/server_school/'
    body = {"server": "123", "school": "456", "e_school": "789"}
    r = requests.post(url=url, data=json.dumps(body), headers=header2)
    r_result = r.content.decode("utf-8")

    # 从数据库取值
    database.data_check()
    data_list = []
    data_result = database.data_check()
    data_list.append(data_result)
    newdata_list = data_list
    create_time = newdata_list[0][0].replace(' ', 'T')

    # 检查点，部分数据来源于数据库
    check_student_info = '{"code":0,"msg":"请求成功","data":{"id":%s,"server":"123","school":"456","e_school":"789"' \
                         ',"create_time":"%s"},"field_name":""}' % (newdata_list[0][1], create_time)
    if check_student_info == r_result:
        print('pass')
    else:
        print('fail')


# case2
def student_school():
    pass


